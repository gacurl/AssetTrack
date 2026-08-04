from __future__ import annotations

import subprocess
from pathlib import Path

from openpyxl import Workbook, load_workbook

from assettrack.import_analysis import analyze_asset_import_xlsx
from scripts.normalize_bq25_shipping_manifest import ASSET_IMPORT_HEADERS, write_normalized_workbook


def _write_manifest(path: Path) -> None:
    workbook = Workbook()
    manifest = workbook.active
    manifest.title = "Manifest A"
    manifest.append(["Intro", "", "", "", "", "", ""])
    manifest.append([None, "Product", "Make\\Model", "Barcode", "Serial", "Case #", "Quantity", "Commnet"])
    manifest.append([None, "Switch", "Cisco 3650CX-8", "DDc4CY002175", "SER-SW-100", "16RU-03", 1, "Preserve first"])
    manifest.append([None, "Laptop", "Dell Latitude", "LAP-100", "SER-LAP-100", "CASE-B", 1, "Excluded laptop"])
    manifest.append([None, "Router", "Cisco ISR4331", "RTR-100", "FCW1931CiPE", "16RU-03", 1, "Preserve second"])
    manifest.append([None, "Product", "Make\\Model", "Barcode", "Seriel", "Case #", "Quantity", "Commnet"])
    manifest.append([None, "Switch", "Cisco 3560CX-8", "SW-101", "SER-SW-101", "16RU-3", 1, "Naming variant"])
    manifest.append([None, "Router", "Cisco ISR4451", "RTR-101", "FOC270rYD9B", "LG-WHE-03", 1, "Reference omitted"])

    other = workbook.create_sheet("Manifest B")
    other.append([None, "Product", "Make", "Model", "Barcode", "Serial", "Case #", "Quantity", "Commnet (F/B)"])
    other.append([None, "Switch", "Cisco", "Catalyst", "SW-102", "FJC27441A0", "LG-WHE-02", 1, "Reference duplicated"])
    other.append([None, "Router", "Cisco", "ASR", "RTR-102", "FOC238X0CG", "LG-WHE-02", 1, "Reference duplicated"])
    other.append([None, "Switch", "Cisco", "Catalyst", "sw-102", "SER-DUP-ASSET", "LG-WHE-02", 1, "Duplicate tag"])
    other.append([None, "Router", "Cisco", "ASR", "RTR-DUP-SERIAL", "FOC238X0CG", "LG-WHE-02", 1, "Duplicate serial"])

    reference = workbook.create_sheet("Reference")
    reference.append(["Case #", "Rack"])
    reference.append(["16RU-03", "Rack 1"])
    reference.append(["16RU-3", "Rack 2"])
    reference.append(["LG-WHE-02", "Rack 3"])
    reference.append(["LG-WHE-02", "Rack 3 duplicate"])

    workbook.save(path)


def _sheet_rows(path: Path, sheet_name: str) -> list[list[object]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook[sheet_name]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def test_shipping_manifest_normalizer_writes_required_sheets_and_asset_import_rows(tmp_path: Path) -> None:
    source = tmp_path / "manifest.xlsx"
    output = tmp_path / "normalized.xlsx"
    _write_manifest(source)

    write_normalized_workbook(source, output)

    workbook = load_workbook(output, data_only=True)
    try:
        assert workbook.sheetnames == ["Asset Import", "Exceptions", "Case Summary", "Read Me"]
    finally:
        workbook.close()

    asset_rows = _sheet_rows(output, "Asset Import")
    assert asset_rows[0] == list(ASSET_IMPORT_HEADERS)
    assert asset_rows[1:] == [
        ["Switch", "DDc4CY002175", "SER-SW-100", "Cisco", "3650CX-8", "16RU-03", "1", "Preserve first"],
        ["Router", "RTR-100", "FCW1931CiPE", "Cisco", "ISR4331", "16RU-03", "2", "Preserve second"],
        ["Switch", "SW-101", "SER-SW-101", "Cisco", "3560CX-8", "16RU-3", "1", "Naming variant"],
        ["Router", "RTR-101", "FOC270rYD9B", "Cisco", "ISR4451", "LG-WHE-03", "1", "Reference omitted"],
        ["Switch", "SW-102", "FJC27441A0", "Cisco", "Catalyst", "LG-WHE-02", "1", "Reference duplicated"],
        ["Router", "RTR-102", "FOC238X0CG", "Cisco", "ASR", "LG-WHE-02", "2", "Reference duplicated"],
    ]

    analysis = analyze_asset_import_xlsx(output, filename="normalized.xlsx")
    assert len(analysis.rows) == 6
    assert analysis.issues == ()


def test_shipping_manifest_normalizer_flags_questionable_values_without_correcting_output(tmp_path: Path) -> None:
    source = tmp_path / "manifest.xlsx"
    output = tmp_path / "normalized.xlsx"
    _write_manifest(source)

    write_normalized_workbook(source, output)

    exceptions = _sheet_rows(output, "Exceptions")
    exception_values = {str(row[6]) for row in exceptions[1:]}
    for value in {"DDc4CY002175", "FCW1931CiPE", "FOC270rYD9B", "FJC27441A0", "FOC238X0CG", "3650CX-8", "3560CX-8"}:
        assert value in exception_values

    exception_reasons = [str(row[7]) for row in exceptions[1:]]
    assert any("Unsupported equipment type" in reason for reason in exception_reasons)
    assert any("Case naming variant also appears in manifest: 16RU-3" in reason for reason in exception_reasons)
    assert any("Case appears in manifest but is omitted from the reference sheet" in reason for reason in exception_reasons)
    assert any("Duplicate case in reference sheet rows" in reason for reason in exception_reasons)

    asset_rows = _sheet_rows(output, "Asset Import")
    flat_output_values = {str(value) for row in asset_rows[1:] for value in row}
    assert "DDc4CY002175" in flat_output_values
    assert "FCW1931CiPE" in flat_output_values
    assert "FOC270rYD9B" in flat_output_values
    assert "FJC27441A0" in flat_output_values
    assert "FOC238X0CG" in flat_output_values
    assert "3650CX-8" in flat_output_values
    assert "3560CX-8" in flat_output_values


def test_shipping_manifest_normalizer_does_not_introduce_duplicate_identifiers(tmp_path: Path) -> None:
    source = tmp_path / "manifest.xlsx"
    output = tmp_path / "normalized.xlsx"
    _write_manifest(source)

    write_normalized_workbook(source, output)

    asset_rows = _sheet_rows(output, "Asset Import")[1:]
    asset_tags = [str(row[1]).upper() for row in asset_rows]
    serial_numbers = [str(row[2]).upper() for row in asset_rows if row[2]]
    assert len(asset_tags) == len(set(asset_tags))
    assert len(serial_numbers) == len(set(serial_numbers))
    assert "SW-102" in asset_tags
    assert "DDC4CY002175" in asset_tags
    assert "RTR-DUP-SERIAL" not in asset_tags

    exceptions = _sheet_rows(output, "Exceptions")
    exception_reasons = [str(row[7]) for row in exceptions[1:]]
    assert any("Duplicate normalized asset tag" in reason for reason in exception_reasons)
    assert any("Duplicate normalized serial number" in reason for reason in exception_reasons)


def test_shipping_manifest_normalizer_output_is_deterministic(tmp_path: Path) -> None:
    source = tmp_path / "manifest.xlsx"
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    _write_manifest(source)

    write_normalized_workbook(source, first)
    write_normalized_workbook(source, second)

    assert first.read_bytes() == second.read_bytes()


def test_shipping_manifest_operational_workbooks_are_not_tracked() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    ignored = subprocess.run(
        ["git", "check-ignore", "data/import/BQ25_shipping_manifest.xlsx"],
        cwd=repo_root,
        check=False,
        capture_output=True,
        text=True,
    )
    assert ignored.returncode == 0

    tracked = subprocess.run(
        ["git", "ls-files", "*BQ25*.xlsx", "*BQ25*.xls", "*BQ25*.csv"],
        cwd=repo_root,
        check=True,
        capture_output=True,
        text=True,
    )
    assert tracked.stdout.strip() == ""
