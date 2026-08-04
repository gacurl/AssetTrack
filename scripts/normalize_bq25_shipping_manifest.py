from __future__ import annotations

import argparse
import warnings
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Iterable

from openpyxl import Workbook, load_workbook

ASSET_IMPORT_HEADERS = (
    "equipment_type",
    "asset_tag",
    "serial_number",
    "manufacturer",
    "model",
    "case_identifier",
    "slot_identifier",
    "notes_comments",
)

EXCEPTION_HEADERS = (
    "source_sheet",
    "source_row",
    "asset_tag",
    "serial_number",
    "case_identifier",
    "field",
    "value",
    "reason",
)

CASE_SUMMARY_HEADERS = (
    "case_identifier",
    "asset_count",
    "slot_start",
    "slot_end",
    "first_source_sheet",
    "first_source_row",
    "last_source_row",
    "reference_status",
)

QUESTIONABLE_VALUES = {
    "DDc4CY002175",
    "FCW1931CiPE",
    "FOC270rYD9B",
    "FJC27441A0",
    "FOC238X0CG",
    "3650CX-8",
    "3560CX-8",
}

HEADER_ALIASES = {
    "equipment_type": {"equipment_type", "equipment type", "type", "device type", "product type", "product"},
    "asset_tag": {"asset_tag", "asset tag", "asset id", "asset", "tag", "barcode"},
    "serial_number": {"serial_number", "serial number", "serial #", "serial no", "serial", "seriel", "s/n", "sn"},
    "manufacturer": {"manufacturer", "mfg", "vendor", "make"},
    "model": {"model", "model number", "product id", "pid", "make\\model", "make/model"},
    "case_identifier": {"case #", "case", "case number", "case no", "case_identifier"},
    "notes_comments": {"notes_comments", "notes/comments", "notes", "comments", "comment", "commnet", "commnet (f/b)"},
}

DEVICE_HEADER_FIELDS = {"equipment_type", "asset_tag", "serial_number", "model"}
SUPPORTED_TYPES = {"switch": "Switch", "router": "Router"}
REFERENCE_SHEET_NAMES = {"name cases"}
FIXED_ZIP_TIMESTAMP = (2026, 1, 1, 0, 0, 0)
FIXED_WORKBOOK_TIMESTAMP = datetime(2026, 1, 1, 0, 0, 0)


@dataclass(frozen=True)
class ManifestRow:
    source_sheet: str
    source_row: int
    equipment_type: str
    asset_tag: str
    serial_number: str
    manufacturer: str
    model: str
    case_identifier: str
    notes_comments: str
    source_order: int


@dataclass(frozen=True)
class ExceptionRow:
    source_sheet: str
    source_row: int
    asset_tag: str
    serial_number: str
    case_identifier: str
    field: str
    value: str
    reason: str


def _text(value: object) -> str:
    return str(value or "").strip()


def _header_key(value: object) -> str:
    return " ".join(_text(value).lower().replace("_", " ").split())


def _canonical_header(value: object) -> str | None:
    key = _header_key(value)
    for canonical, aliases in HEADER_ALIASES.items():
        if key in aliases:
            return canonical
    return None


def _header_map(values: Iterable[object]) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, value in enumerate(values):
        canonical = _canonical_header(value)
        if canonical is not None and canonical not in headers:
            headers[canonical] = index
    return headers


def _is_manifest_header(headers: dict[str, int]) -> bool:
    return "case_identifier" in headers and bool(DEVICE_HEADER_FIELDS & set(headers))


def _is_reference_header(headers: dict[str, int]) -> bool:
    return "case_identifier" in headers and not bool(DEVICE_HEADER_FIELDS & set(headers))


def _cell(row: tuple[object, ...], headers: dict[str, int], field: str) -> str:
    index = headers.get(field)
    if index is None or index >= len(row):
        return ""
    return _text(row[index])


def _normalize_identity(value: str) -> str:
    return value.strip().upper()


def _load_workbook_read_only(path: Path):
    return load_workbook(path, read_only=True, data_only=True)


def _maybe_case_variant(case_identifier: str, cases: set[str]) -> str | None:
    if "-" not in case_identifier:
        return None
    prefix, suffix = case_identifier.rsplit("-", 1)
    if not suffix.isdigit():
        return None
    variant = f"{prefix}-{int(suffix)}"
    if variant != case_identifier and variant in cases:
        return variant
    padded_variant = f"{prefix}-{int(suffix):02d}"
    if padded_variant != case_identifier and padded_variant in cases:
        return padded_variant
    return None


def _split_make_model(value: str) -> tuple[str, str]:
    parts = value.split(None, 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return "", value


def _append_exception(
    exceptions: list[ExceptionRow],
    row: ManifestRow,
    *,
    field: str,
    value: str,
    reason: str,
) -> None:
    exceptions.append(
        ExceptionRow(
            source_sheet=row.source_sheet,
            source_row=row.source_row,
            asset_tag=row.asset_tag,
            serial_number=row.serial_number,
            case_identifier=row.case_identifier,
            field=field,
            value=value,
            reason=reason,
        )
    )


def read_manifest(path: Path) -> tuple[list[ManifestRow], list[ExceptionRow], dict[str, list[int]]]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Cannot parse header or footer so it will be ignored",
            category=UserWarning,
        )
        workbook = _load_workbook_read_only(path)
        rows: list[ManifestRow] = []
        exceptions: list[ExceptionRow] = []
        reference_cases: dict[str, list[int]] = defaultdict(list)
        source_order = 0

        try:
            for worksheet in workbook.worksheets:
                if _header_key(worksheet.title) in REFERENCE_SHEET_NAMES:
                    for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
                        for value in values:
                            case_identifier = _text(value).split(" - ", 1)[0]
                            if case_identifier and not isinstance(value, datetime):
                                reference_cases[case_identifier].append(row_number)
                    continue

                active_headers: dict[str, int] | None = None
                reference_headers: dict[str, int] | None = None
                for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    headers = _header_map(values)
                    if _is_manifest_header(headers):
                        active_headers = headers
                        reference_headers = None
                        continue
                    if _is_reference_header(headers):
                        reference_headers = headers
                        active_headers = None
                        continue

                    if active_headers is not None:
                        case_identifier = _cell(values, active_headers, "case_identifier")
                        equipment_type = _cell(values, active_headers, "equipment_type")
                        asset_tag = _cell(values, active_headers, "asset_tag")
                        serial_number = _cell(values, active_headers, "serial_number")
                        model = _cell(values, active_headers, "model")
                        manufacturer = _cell(values, active_headers, "manufacturer")
                        if not manufacturer and model:
                            manufacturer, model = _split_make_model(model)
                        notes_comments = _cell(values, active_headers, "notes_comments")
                        if not any((case_identifier, equipment_type, asset_tag, serial_number, model, manufacturer, notes_comments)):
                            continue
                        source_order += 1
                        manifest_row = ManifestRow(
                            source_sheet=worksheet.title,
                            source_row=row_number,
                            equipment_type=equipment_type,
                            asset_tag=asset_tag,
                            serial_number=serial_number,
                            manufacturer=manufacturer,
                            model=model,
                            case_identifier=case_identifier,
                            notes_comments=notes_comments,
                            source_order=source_order,
                        )
                        if not case_identifier:
                            _append_exception(
                                exceptions,
                                manifest_row,
                                field="Case #",
                                value=case_identifier,
                                reason="Missing Case #; row excluded from Asset Import.",
                            )
                            continue
                        rows.append(manifest_row)
                        continue

                    if reference_headers is not None:
                        case_identifier = _cell(values, reference_headers, "case_identifier")
                        if case_identifier:
                            reference_cases[case_identifier].append(row_number)
        finally:
            workbook.close()

    return rows, exceptions, dict(reference_cases)


def normalize_manifest(path: Path) -> tuple[list[list[str]], list[ExceptionRow], list[list[object]]]:
    source_rows, exceptions, reference_cases = read_manifest(path)
    included: list[ManifestRow] = []
    seen_asset_tags: dict[str, ManifestRow] = {}
    seen_serial_numbers: dict[str, ManifestRow] = {}

    for row in source_rows:
        normalized_type = row.equipment_type.strip().lower()
        if normalized_type not in SUPPORTED_TYPES:
            _append_exception(
                exceptions,
                row,
                field="equipment_type",
                value=row.equipment_type,
                reason="Unsupported equipment type; only Switch and Router rows are included.",
            )
            continue

        normalized_asset_tag = _normalize_identity(row.asset_tag)
        if not normalized_asset_tag:
            _append_exception(
                exceptions,
                row,
                field="asset_tag",
                value=row.asset_tag,
                reason="Missing asset tag; row excluded from Asset Import.",
            )
            continue
        previous_asset = seen_asset_tags.get(normalized_asset_tag)
        if previous_asset is not None:
            _append_exception(
                exceptions,
                row,
                field="asset_tag",
                value=row.asset_tag,
                reason=f"Duplicate normalized asset tag; conflicts with {previous_asset.source_sheet} row {previous_asset.source_row}.",
            )
            continue

        normalized_serial = _normalize_identity(row.serial_number)
        if normalized_serial:
            previous_serial = seen_serial_numbers.get(normalized_serial)
            if previous_serial is not None:
                _append_exception(
                    exceptions,
                    row,
                    field="serial_number",
                    value=row.serial_number,
                    reason=f"Duplicate normalized serial number; conflicts with {previous_serial.source_sheet} row {previous_serial.source_row}.",
                )
                continue
            seen_serial_numbers[normalized_serial] = row

        seen_asset_tags[normalized_asset_tag] = row
        included.append(row)

        for field, value in (
            ("asset_tag", row.asset_tag),
            ("serial_number", row.serial_number),
            ("model", row.model),
            ("case_identifier", row.case_identifier),
        ):
            if value in QUESTIONABLE_VALUES:
                _append_exception(
                    exceptions,
                    row,
                    field=field,
                    value=value,
                    reason="Known questionable source value; preserved unchanged.",
                )

    manifest_cases = {row.case_identifier for row in included}
    for row in included:
        variant = _maybe_case_variant(row.case_identifier, manifest_cases)
        if variant is not None:
            _append_exception(
                exceptions,
                row,
                field="Case #",
                value=row.case_identifier,
                reason=f"Case naming variant also appears in manifest: {variant}; preserved unchanged.",
            )
        if reference_cases and row.case_identifier not in reference_cases:
            _append_exception(
                exceptions,
                row,
                field="Case #",
                value=row.case_identifier,
                reason="Case appears in manifest but is omitted from the reference sheet; preserved unchanged.",
            )

    for case_identifier, rows in sorted(reference_cases.items()):
        if len(rows) > 1:
            exceptions.append(
                ExceptionRow(
                    source_sheet="Reference",
                    source_row=rows[0],
                    asset_tag="",
                    serial_number="",
                    case_identifier=case_identifier,
                    field="Case #",
                    value=case_identifier,
                    reason=f"Duplicate case in reference sheet rows: {', '.join(str(row_number) for row_number in rows)}.",
                )
            )

    rows_by_case: dict[str, list[ManifestRow]] = defaultdict(list)
    for row in included:
        rows_by_case[row.case_identifier].append(row)

    asset_rows: list[list[str]] = []
    case_summary_rows: list[list[object]] = []
    for case_identifier, case_rows in rows_by_case.items():
        case_rows.sort(key=lambda row: row.source_order)
        for slot_index, row in enumerate(case_rows, start=1):
            asset_rows.append(
                [
                    SUPPORTED_TYPES[row.equipment_type.strip().lower()],
                    row.asset_tag,
                    row.serial_number,
                    row.manufacturer,
                    row.model,
                    row.case_identifier,
                    str(slot_index),
                    row.notes_comments,
                ]
            )
        reference_status = "not checked"
        if reference_cases:
            reference_status = "present" if case_identifier in reference_cases else "missing from reference"
        case_summary_rows.append(
            [
                case_identifier,
                len(case_rows),
                1,
                len(case_rows),
                case_rows[0].source_sheet,
                case_rows[0].source_row,
                case_rows[-1].source_row,
                reference_status,
            ]
        )

    return asset_rows, exceptions, case_summary_rows


def _append_rows(worksheet, headers: tuple[str, ...], rows: Iterable[Iterable[object]]) -> None:
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))


def build_output_workbook(path: Path) -> Workbook:
    asset_rows, exceptions, case_summary_rows = normalize_manifest(path)
    workbook = Workbook()
    workbook.properties.created = FIXED_WORKBOOK_TIMESTAMP
    workbook.properties.modified = FIXED_WORKBOOK_TIMESTAMP

    asset_sheet = workbook.active
    asset_sheet.title = "Asset Import"
    _append_rows(asset_sheet, ASSET_IMPORT_HEADERS, asset_rows)

    exceptions_sheet = workbook.create_sheet("Exceptions")
    _append_rows(
        exceptions_sheet,
        EXCEPTION_HEADERS,
        (
            [
                exception.source_sheet,
                exception.source_row,
                exception.asset_tag,
                exception.serial_number,
                exception.case_identifier,
                exception.field,
                exception.value,
                exception.reason,
            ]
            for exception in exceptions
        ),
    )

    case_summary_sheet = workbook.create_sheet("Case Summary")
    _append_rows(case_summary_sheet, CASE_SUMMARY_HEADERS, case_summary_rows)

    read_me_sheet = workbook.create_sheet("Read Me")
    _append_rows(
        read_me_sheet,
        ("item", "value"),
        (
            ("Workflow", "Use Admin -> Import Assets at /admin/assets/import with the Asset Import sheet."),
            ("Database writes", "This normalizer does not connect to or write the AssetTrack database."),
            ("Storage source", "Case # is preserved as case_identifier. Slots are assigned by manifest order within each case."),
            ("Supported rows", "Only Switch and Router rows are included in Asset Import."),
            ("Exceptions", "Questionable, unsupported, duplicate, or reference-mismatch values are listed without correction."),
        ),
    )
    return workbook


def _save_workbook_deterministically(workbook: Workbook, output_path: Path) -> None:
    with NamedTemporaryFile(suffix=".xlsx") as handle:
        workbook.save(handle.name)
        handle.seek(0)
        source_bytes = handle.read()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(BytesIO(source_bytes), "r") as source_zip:
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as output_zip:
            for name in sorted(source_zip.namelist()):
                source_info = source_zip.getinfo(name)
                output_info = zipfile.ZipInfo(filename=name, date_time=FIXED_ZIP_TIMESTAMP)
                output_info.compress_type = zipfile.ZIP_DEFLATED
                output_info.external_attr = source_info.external_attr
                output_zip.writestr(output_info, source_zip.read(name))


def write_normalized_workbook(input_path: Path, output_path: Path) -> None:
    workbook = build_output_workbook(input_path)
    _save_workbook_deterministically(workbook, output_path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Normalize a BQ25 shipping manifest for AssetTrack Import Assets.")
    parser.add_argument("input", type=Path, help="Source BQ25 shipping manifest workbook.")
    parser.add_argument("output", type=Path, help="Normalized AssetTrack-compatible workbook to write.")
    args = parser.parse_args(argv)

    write_normalized_workbook(args.input, args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
